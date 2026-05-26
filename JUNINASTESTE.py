import streamlit as st
import pandas as pd
import plotly.express as px
from datetime import datetime
import unicodedata
import re
from io import BytesIO

# =========================================================
# FALLBACK OPENPYXL
# =========================================================

try:
    from openpyxl.styles import PatternFill, Font, Alignment
    OPENPYXL_DISPONIVEL = True
except ModuleNotFoundError:
    OPENPYXL_DISPONIVEL = False

# =========================================================
# CONFIGURAÇÃO DA PÁGINA
# =========================================================

st.set_page_config(
    page_title="OPERAÇÃO SÃO JOÃO 2026",
    page_icon="🚔",
    layout="wide"
)

# =========================================================
# FORÇAR IDIOMA PORTUGUÊS NO FRONT
# =========================================================

st.markdown("""
<script>
document.documentElement.lang = "pt-BR";
</script>
""", unsafe_allow_html=True)


# =========================================================
# HEADER PROFISSIONAL REAL (SEM BUG)
# =========================================================

col1, col2 = st.columns([1, 25])

with col1:
    st.image("brasao.png", width=60)

with col2:
    st.markdown("""
        <div style="
            background: linear-gradient(90deg, #0B1F3A 0%, #1E3A5F 100%);
            padding: 12px 20px;
            border-radius: 8px;
            margin-left: -12px;
            height: 65px;
            display: flex;
            align-items: center;
        ">
            <span style="
                color: white;
                font-size: 2rem;
                font-weight: 800;
                letter-spacing: 1px;
                line-height: 1;
            ">
                OPERAÇÃO SÃO JOÃO 2026 SUBCHEFIA DE OPERAÇÕES PMDF
            </span>
        </div>
    """, unsafe_allow_html=True)

# =========================================================
# LINK DA PLANILHA
# =========================================================

url = (
    "https://docs.google.com/spreadsheets/d/"
    "1U2re0vGfssfUAFzra2oJoIVp4klRsbkpWBXVXSFd2Rs/"
    "export?format=csv&gid=459543687"
)

# =========================================================
# CORES
# =========================================================

PALETA_BARRAS = [
    "#5B8FF9", "#61DDAA", "#65789B", "#F6BD16",
    "#7262FD", "#78D3F8", "#9661BC", "#F6903D",
    "#008685", "#F08BB4"
]

PALETA_PIZZA = [
    "#5B8FF9", "#61DDAA", "#F6BD16", "#7262FD",
    "#F08BB4", "#78D3F8", "#65789B", "#F6903D",
    "#008685", "#9661BC"
]

COR_PADRAO_BARRA = "#5B8FF9"
COR_LINHA_2 = "#61DDAA"
COR_HISTOGRAMA = "#9CC3FF"

MAPA_MESES = {
    1: "Jan", 2: "Fev", 3: "Mar", 4: "Abr", 5: "Mai", 6: "Jun",
    7: "Jul", 8: "Ago", 9: "Set", 10: "Out", 11: "Nov", 12: "Dez"
}

DIAS_SEMANA = ["Seg", "Ter", "Qua", "Qui", "Sex", "Sáb", "Dom"]

# =========================================================
# FUNÇÕES
# =========================================================

def normalizar_texto(texto):
    texto = str(texto).strip().upper()
    texto = unicodedata.normalize("NFKD", texto).encode("ASCII", "ignore").decode("utf-8")
    texto = re.sub(r"\s+", " ", texto)
    return texto

def limpar_categoria(valor):
    if pd.isna(valor):
        return pd.NA
    valor = str(valor).strip()
    if valor == "":
        return pd.NA
    valor = re.sub(r"\s+", " ", valor)
    return valor.upper()

def localizar_coluna(colunas, termos):
    for termo in termos:
        termo_norm = normalizar_texto(termo)
        for coluna in colunas:
            if termo_norm == normalizar_texto(coluna):
                return coluna

    for termo in termos:
        termo_norm = normalizar_texto(termo)
        for coluna in colunas:
            if termo_norm in normalizar_texto(coluna):
                return coluna
    return None

def converter_numero_misto(valor):
    if pd.isna(valor):
        return pd.NA

    s = str(valor).strip()
    if s == "":
        return pd.NA

    s = s.replace(" ", "")
    s = re.sub(r"[R$\u00A0]", "", s)

    if "," in s and "." in s:
        if s.rfind(",") > s.rfind("."):
            s = s.replace(".", "")
            s = s.replace(",", ".")
        else:
            s = s.replace(",", "")
    elif "," in s:
        s = s.replace(".", "")
        s = s.replace(",", ".")
    else:
        s = s.replace(",", "")

    s = re.sub(r"[^0-9\.-]", "", s)

    try:
        return float(s)
    except:
        return pd.NA

def parse_data_strict(valor):
    if pd.isna(valor):
        return pd.NaT

    s = str(valor).strip()
    if s in ["", "nan", "None", "0", "0000-00-00", "0000-00-00 00:00:00", "00/00/0000"]:
        return pd.NaT

    s = re.sub(r"\s+", " ", s)

    formatos = [
        "%d/%m/%Y",
        "%d/%m/%Y %H:%M",
        "%d/%m/%Y %H:%M:%S",
        "%Y-%m-%d",
        "%Y-%m-%d %H:%M",
        "%Y-%m-%d %H:%M:%S",
    ]

    for fmt in formatos:
        try:
            return pd.Timestamp(datetime.strptime(s, fmt))
        except:
            pass

    return pd.NaT

def aplicar_estilo(fig):
    fig.update_layout(
        template="simple_white",
        plot_bgcolor="white",
        paper_bgcolor="white",
        font=dict(size=13, color="#3B4A5A"),
        bargap=0.22,
        margin=dict(l=20, r=20, t=60, b=30),
        xaxis=dict(
            showgrid=False,
            zeroline=False,
            linecolor="#D9E2EC",
            tickfont=dict(color="#52606D")
        ),
        yaxis=dict(
            showgrid=True,
            gridcolor="#EEF2F7",
            zeroline=False,
            linecolor="#D9E2EC",
            tickfont=dict(color="#52606D")
        ),
        legend=dict(
            bgcolor="rgba(0,0,0,0)",
            orientation="h",
            yanchor="bottom",
            y=1.02,
            xanchor="right",
            x=1
        )
    )
    return fig

@st.cache_data(ttl=60)
def carregar_dados():
    df = pd.read_csv(url)
    df = df.dropna(how="all")
    df.columns = [str(c).strip() for c in df.columns]
    return df

def obter_cor_por_upm(valor_upm, texto_row=""):
    upm = str(valor_upm).strip().upper()

    if "EVENTO CANCELADO" in str(texto_row).upper():
        return "#F8D7DA"

    if "BPESC" in upm or "BPESCOLAR" in upm:
        return "#D9EAF7"

    if "BPRURAL" in upm or upm in ["BPR", "BPRV"]:
        return "#DFF0D8"

    if "BPMA" in upm:
        return "#E6D9F2"

    if upm != "":
        return "#FFF3CD"

    return ""

def definir_cor_linha(row, coluna_upm):
    valor_upm = row[coluna_upm] if coluna_upm in row.index else ""
    texto_row = " ".join([str(v).strip() for v in row.fillna("").tolist()])

    cor = obter_cor_por_upm(valor_upm, texto_row)

    if cor:
        estilo = f"background-color: {cor}; color: #111827; font-weight: bold;"
        return [estilo] * len(row)

    return ["color: #111827;"] * len(row)

def gerar_excel_colorido(df_exportacao, coluna_upm):
    if not OPENPYXL_DISPONIVEL:
        return None

    output = BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df_exportacao.to_excel(writer, index=False, sheet_name="DADOS_2026")

        worksheet = writer.sheets["DADOS_2026"]

        fill_azul = PatternFill(fill_type="solid", start_color="D9EAF7", end_color="D9EAF7")
        fill_verde = PatternFill(fill_type="solid", start_color="DFF0D8", end_color="DFF0D8")
        fill_roxo = PatternFill(fill_type="solid", start_color="E6D9F2", end_color="E6D9F2")
        fill_vermelho = PatternFill(fill_type="solid", start_color="F8D7DA", end_color="F8D7DA")
        fill_amarelo = PatternFill(fill_type="solid", start_color="FFF3CD", end_color="FFF3CD")
        fill_header = PatternFill(fill_type="solid", start_color="D9E2EC", end_color="D9E2EC")

        for cell in worksheet[1]:
            cell.font = Font(bold=True)
            cell.fill = fill_header
            cell.alignment = Alignment(horizontal="center", vertical="center")

        idx_upm_excel = None
        for idx, col_name in enumerate(df_exportacao.columns, start=1):
            if str(col_name).strip() == str(coluna_upm).strip():
                idx_upm_excel = idx
                break

        for row_idx in range(2, worksheet.max_row + 1):
            valores = []
            for col_idx in range(1, worksheet.max_column + 1):
                valor = worksheet.cell(row=row_idx, column=col_idx).value
                valores.append("" if valor is None else str(valor).strip())

            texto_row = " ".join(valores)
            valor_upm = ""

            if idx_upm_excel is not None:
                valor_upm = worksheet.cell(row=row_idx, column=idx_upm_excel).value
                valor_upm = "" if valor_upm is None else str(valor_upm).strip()

            cor = obter_cor_por_upm(valor_upm, texto_row)

            if cor == "#D9EAF7":
                fill = fill_azul
            elif cor == "#DFF0D8":
                fill = fill_verde
            elif cor == "#E6D9F2":
                fill = fill_roxo
            elif cor == "#F8D7DA":
                fill = fill_vermelho
            elif cor == "#FFF3CD":
                fill = fill_amarelo
            else:
                fill = None

            if fill is not None:
                for col_idx in range(1, worksheet.max_column + 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    cell.fill = fill
                    cell.font = Font(bold=True, color="111827")

        for col in worksheet.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    valor = "" if cell.value is None else str(cell.value)
                    if len(valor) > max_length:
                        max_length = len(valor)
                except:
                    pass
            worksheet.column_dimensions[col_letter].width = min(max_length + 2, 40)

    output.seek(0)
    return output.getvalue()

# =========================================================
# INÍCIO
# =========================================================

try:
    df = carregar_dados()
    df_original = df.copy()

    horario = datetime.now().strftime("%d/%m/%Y %H:%M:%S")

    st.sidebar.success("🟢 DASHBOARD SINCRONIZADO")
    st.sidebar.info(f"ÚLTIMA ATUALIZAÇÃO:\n{horario}")

    if not OPENPYXL_DISPONIVEL:
        st.warning("Openpyxl não instalado no ambiente. O download em Excel colorido ficará desativado.")

    colunas = df.columns.tolist()

    coluna_cpr = df.columns[2]
    coluna_upm = df.columns[3]
    coluna_inicio = df.columns[9]
    coluna_fim = df.columns[10]

    coluna_comando = localizar_coluna(colunas, ["COMANDO"])
    coluna_cidade = localizar_coluna(colunas, ["CIDADE", "MUNICIPIO"])
    coluna_evento = localizar_coluna(colunas, ["EVENTO", "NOME EVENTO"])
    coluna_publico = localizar_coluna(colunas, ["PUBLICO", "PÚBLICO", "PUBLICO PREVISTO", "PÚBLICO PREVISTO"])
    coluna_natureza = localizar_coluna(colunas, ["NATUREZA"])
    coluna_tipo_publico = localizar_coluna(colunas, ["TIPO DE PUBLICO", "TIPO DE PÚBLICO", "TIPO PUBLICO"])

    if coluna_publico:
        df[coluna_publico] = df[coluna_publico].apply(converter_numero_misto)
        df[coluna_publico] = pd.to_numeric(df[coluna_publico], errors="coerce")

    df[coluna_inicio] = df[coluna_inicio].apply(parse_data_strict)
    df[coluna_fim] = df[coluna_fim].apply(parse_data_strict)

    if coluna_natureza:
        df[coluna_natureza] = df[coluna_natureza].apply(limpar_categoria)

    if coluna_tipo_publico:
        df[coluna_tipo_publico] = df[coluna_tipo_publico].apply(limpar_categoria)

    if coluna_comando:
        df[coluna_comando] = df[coluna_comando].astype(str).str.strip()

    if coluna_cidade:
        df[coluna_cidade] = df[coluna_cidade].astype(str).str.strip()

    if coluna_evento:
        df[coluna_evento] = df[coluna_evento].astype(str).str.strip()

    if coluna_cpr:
        df[coluna_cpr] = df[coluna_cpr].astype(str).str.strip()

    if coluna_upm:
        df[coluna_upm] = df[coluna_upm].astype(str).str.strip()

    df["_ID_LINHA_EVENTO_"] = range(1, len(df) + 1)

    df["DATA_INICIO_BASE"] = df[coluna_inicio]
    df["DATA_FIM_BASE"] = df[coluna_fim]
    df["DATA_EVENTO_BASE"] = df[coluna_inicio]

    total_linhas_original = len(df_original)

    df = df[df["DATA_INICIO_BASE"].notna()].copy()
    df["DATA_FIM_BASE"] = df["DATA_FIM_BASE"].fillna(df["DATA_INICIO_BASE"])
    df.loc[df["DATA_FIM_BASE"] < df["DATA_INICIO_BASE"], "DATA_FIM_BASE"] = df["DATA_INICIO_BASE"]

    df["Ano"] = df["DATA_EVENTO_BASE"].dt.year.astype("Int64")
    df["Mes_Num"] = df["DATA_EVENTO_BASE"].dt.month.astype("Int64")
    df["Mes_Abrev"] = df["Mes_Num"].map(MAPA_MESES)
    df["Hora"] = df["DATA_INICIO_BASE"].dt.hour.astype("Int64")

    df = df[df["Ano"].notna()].copy()
    df = df[df["Ano"] != 2022].copy()

    # =====================================================
    # FILTROS
    # =====================================================

    df_filtrado = df.copy()

    st.sidebar.subheader("🎯 FILTROS")

    anos = sorted(df_filtrado["Ano"].dropna().astype(int).unique().tolist())
    anos_sel = st.sidebar.multiselect("FILTRAR ANO", options=anos, default=[2026] if 2026 in anos else [])
    if anos_sel:
        df_filtrado = df_filtrado[df_filtrado["Ano"].isin(anos_sel)]

    if coluna_cpr:
        cpr_opcoes = sorted(df_filtrado[coluna_cpr].dropna().astype(str).unique().tolist())
        cpr_sel = st.sidebar.multiselect("FILTRAR CPR", options=cpr_opcoes, default=[])
        if cpr_sel:
            df_filtrado = df_filtrado[df_filtrado[coluna_cpr].astype(str).isin(cpr_sel)]

    if coluna_upm:
        upm_opcoes = sorted(df_filtrado[coluna_upm].dropna().astype(str).unique().tolist())
        upm_sel = st.sidebar.multiselect("FILTRAR UPM", options=upm_opcoes, default=[])
        if upm_sel:
            df_filtrado = df_filtrado[df_filtrado[coluna_upm].astype(str).isin(upm_sel)]

    if coluna_comando:
        comandos = sorted(df_filtrado[coluna_comando].dropna().astype(str).unique().tolist())
        comandos_sel = st.sidebar.multiselect("FILTRAR COMANDO", options=comandos, default=[])
        if comandos_sel:
            df_filtrado = df_filtrado[df_filtrado[coluna_comando].astype(str).isin(comandos_sel)]

    if coluna_cidade:
        cidades = sorted(df_filtrado[coluna_cidade].dropna().astype(str).unique().tolist())
        cidades_sel = st.sidebar.multiselect("FILTRAR CIDADE", options=cidades, default=[])
        if cidades_sel:
            df_filtrado = df_filtrado[df_filtrado[coluna_cidade].astype(str).isin(cidades_sel)]

    if coluna_natureza:
        natureza = sorted(df_filtrado[coluna_natureza].dropna().astype(str).unique().tolist())
        natureza_sel = st.sidebar.multiselect("FILTRAR NATUREZA", options=natureza, default=[])
        if natureza_sel:
            df_filtrado = df_filtrado[df_filtrado[coluna_natureza].astype(str).isin(natureza_sel)]

    if df_filtrado["DATA_INICIO_BASE"].notna().any() and df_filtrado["DATA_FIM_BASE"].notna().any():

        intervalo = st.sidebar.date_input(
            "FILTRAR PERÍODO",
            value=(),
            format="DD/MM/YYYY"
        )

        if isinstance(intervalo, (tuple, list)) and len(intervalo) == 2:
            data_ini, data_fim = intervalo

            df_filtrado = df_filtrado[
                (df_filtrado["DATA_INICIO_BASE"].dt.date <= data_fim) &
                (df_filtrado["DATA_FIM_BASE"].dt.date >= data_ini)
                ]

    if df_filtrado.empty:
        st.warning("NENHUM REGISTRO ENCONTRADO.")
        st.stop()

    df_2026 = df_filtrado[df_filtrado["Ano"] == 2026].copy()

    if df_2026.empty:
        st.warning("NÃO HÁ DADOS DE 2026 PARA EXIBIR OS GRÁFICOS OPERACIONAIS.")
        st.stop()

    # =====================================================
    # CONFERÊNCIA DOS DADOS
    # =====================================================

    with st.expander("🔎 CONFERÊNCIA DOS DADOS", expanded=False):
        resumo_ano = (
            df.groupby("Ano")["_ID_LINHA_EVENTO_"]
            .count()
            .reset_index(name="Eventos")
            .sort_values("Ano")
        )

        resumo_meses = (
            df.groupby(["Ano", "Mes_Num", "Mes_Abrev"])["_ID_LINHA_EVENTO_"]
            .count()
            .reset_index(name="Eventos")
            .sort_values(["Ano", "Mes_Num"])
        )

        eventos_2026 = int(df[df["Ano"] == 2026]["_ID_LINHA_EVENTO_"].count())
        eventos_jan = int(df[df["Mes_Num"] == 1]["_ID_LINHA_EVENTO_"].count())
        eventos_fev = int(df[df["Mes_Num"] == 2]["_ID_LINHA_EVENTO_"].count())
        descartadas = total_linhas_original - len(df)

        c1, c2, c3, c4 = st.columns(4)
        c1.metric("LINHAS ORIGINAIS", total_linhas_original)
        c2.metric("LINHAS VÁLIDAS EM J", len(df))
        c3.metric("EVENTOS 2026", eventos_2026)
        c4.metric("DESCARTADAS", int(descartadas))

        c5, c6 = st.columns(2)
        c5.metric("EVENTOS EM JANEIRO", eventos_jan)
        c6.metric("EVENTOS EM FEVEREIRO", eventos_fev)

        st.caption(f"Coluna J usada como início: {coluna_inicio}")
        st.caption(f"Coluna K usada como fim: {coluna_fim}")

        st.markdown("### Eventos por ano")
        st.dataframe(resumo_ano, use_container_width=True)

        st.markdown("### Eventos por mês")
        st.dataframe(resumo_meses, use_container_width=True)

    # =====================================================
    # PANORAMA GERAL
    # =====================================================

    st.subheader("📚 PANORAMA GERAL")

    st.markdown("### 📊 COMPARATIVO DE EVENTOS POR ANO")
    eventos_ano = (
        df.groupby("Ano")["_ID_LINHA_EVENTO_"]
        .count()
        .reset_index(name="Eventos")
        .sort_values("Ano")
    )

    eventos_ano["Ano"] = eventos_ano["Ano"].astype(int).astype(str)
    eventos_ano["Eventos_txt"] = eventos_ano["Eventos"].astype(int).astype(str)

    fig_ano = px.bar(
        eventos_ano,
        x="Ano",
        y="Eventos",
        text="Eventos_txt",
        color="Ano",
        color_discrete_sequence=PALETA_BARRAS
    )

    fig_ano.update_traces(textposition="outside")
    fig_ano.update_xaxes(type="category", categoryorder="category ascending")
    fig_ano.update_yaxes(tickformat=",d")
    fig_ano = aplicar_estilo(fig_ano)
    st.plotly_chart(fig_ano, use_container_width=True, config={"locale": "pt-BR"})

    st.markdown("### 📅 COMPARATIVO ANO/MÊS")
    eventos_mes = (
        df.groupby(["Ano", "Mes_Num", "Mes_Abrev"])["_ID_LINHA_EVENTO_"]
        .count()
        .reset_index(name="Eventos")
        .sort_values(["Ano", "Mes_Num"])
    )

    eventos_mes = eventos_mes[eventos_mes["Eventos"] > 0].copy()
    eventos_mes = eventos_mes[eventos_mes["Mes_Num"].notna()].copy()

    if not eventos_mes.empty:
        meses_existentes = sorted(eventos_mes["Mes_Num"].astype(int).unique().tolist())
        ordem_meses_existentes = [MAPA_MESES[m] for m in meses_existentes if m in MAPA_MESES]

        eventos_mes["Mes_Abrev"] = pd.Categorical(
            eventos_mes["Mes_Abrev"],
            categories=ordem_meses_existentes,
            ordered=True
        )
        eventos_mes["Ano"] = eventos_mes["Ano"].astype(int).astype(str)

        fig_mes = px.bar(
            eventos_mes,
            x="Mes_Abrev",
            y="Eventos",
            color="Ano",
            barmode="group",
            text_auto=True,
            color_discrete_sequence=PALETA_BARRAS
        )
        fig_mes.update_xaxes(categoryorder="array", categoryarray=ordem_meses_existentes)
        fig_mes.update_yaxes(tickformat=",d")
        fig_mes = aplicar_estilo(fig_mes)
        st.plotly_chart(fig_mes, use_container_width=True, config={"locale": "pt-BR"})

        # =====================================================
        # INDICADORES
        # =====================================================

    st.subheader("📌 INDICADORES OPERACIONAIS")

    total_eventos = int(df_2026["_ID_LINHA_EVENTO_"].count())
    total_publico = int(df_2026[coluna_publico].fillna(0).sum()) if coluna_publico else 0
    total_cidades = df_2026[coluna_cidade].dropna().nunique() if coluna_cidade else 0
    total_cprs = df_2026[coluna_cpr].dropna().nunique() if coluna_cpr else 0
    total_upms = df_2026[coluna_upm].dropna().nunique() if coluna_upm else 0

    c1, c2, c3, c4, c5 = st.columns(5)
    c1.metric("🎉 EVENTOS", total_eventos)
    c2.metric("👥 PÚBLICO", f"{total_publico:,}".replace(",", "."))
    c3.metric("🏙️ CIDADES", total_cidades)
    c4.metric("🛡️ CPR", total_cprs)
    c5.metric("🏢 UPM", total_upms)

    # =====================================================
    # MAPA GEOGRÁFICO DOS EVENTOS - DISTRITO FEDERAL
    # =====================================================

    st.subheader("🗺️ MAPA GEOGRÁFICO DOS EVENTOS - DISTRITO FEDERAL")

    coluna_lat = localizar_coluna(colunas, ["LATITUDE", "LAT"])
    coluna_lon = localizar_coluna(colunas, ["LONGITUDE", "LON", "LONG"])

    if coluna_lat and coluna_lon:
        mapa_df = df_2026.copy()

        mapa_df[coluna_lat] = (
            mapa_df[coluna_lat]
            .astype(str)
            .str.strip()
            .str.replace(",", ".", regex=False)
        )

        mapa_df[coluna_lon] = (
            mapa_df[coluna_lon]
            .astype(str)
            .str.strip()
            .str.replace(",", ".", regex=False)
        )

        mapa_df[coluna_lat] = pd.to_numeric(mapa_df[coluna_lat], errors="coerce")
        mapa_df[coluna_lon] = pd.to_numeric(mapa_df[coluna_lon], errors="coerce")

        mapa_df = mapa_df[
            mapa_df[coluna_lat].notna() &
            mapa_df[coluna_lon].notna()
        ].copy()

        # recorte aproximado do Distrito Federal
        mapa_df = mapa_df[
            mapa_df[coluna_lat].between(-16.10, -15.45) &
            mapa_df[coluna_lon].between(-48.30, -47.25)
        ].copy()

        if not mapa_df.empty:
            hover_cols = {}
            if coluna_cidade:
                hover_cols[coluna_cidade] = True
            if coluna_cpr:
                hover_cols[coluna_cpr] = True
            if coluna_upm:
                hover_cols[coluna_upm] = True
            if coluna_publico:
                hover_cols[coluna_publico] = True

            fig_mapa = px.scatter_map(
                mapa_df,
                lat=coluna_lat,
                lon=coluna_lon,
                hover_name=coluna_evento if coluna_evento else coluna_cidade,
                hover_data=hover_cols,
                color=coluna_cpr if coluna_cpr else None,
                size=coluna_publico if coluna_publico else None,
                zoom=9.5,
                center={"lat": -15.79, "lon": -47.88},
                height=550
            )

            fig_mapa.update_traces(marker=dict(opacity=0.70))
            fig_mapa.update_layout(
                margin=dict(l=10, r=10, t=50, b=10)
            )

            st.plotly_chart(
                fig_mapa,
                use_container_width=True,
                config={"locale": "pt-BR"}
            )
        else:
            st.warning("Não há eventos com latitude e longitude válidas dentro da área do Distrito Federal.")
    else:
        st.warning("As colunas de latitude e longitude não foram localizadas na planilha.")
    # =====================================================
    # EVENTOS POR CPR - FIXO 2026
    # =====================================================

    if coluna_cpr:
        st.subheader("🛡️ EVENTOS POR CPR")

        cpr_df = (
            df_2026[df_2026[coluna_cpr].notna()]
            .groupby(coluna_cpr)["_ID_LINHA_EVENTO_"]
            .count()
            .reset_index(name="Eventos")
            .sort_values(by="Eventos", ascending=True)
        )

        if not cpr_df.empty:
            cpr_df["categoria_cor"] = cpr_df[coluna_cpr].astype(str)

            fig_cpr = px.bar(
                cpr_df,
                x="Eventos",
                y=coluna_cpr,
                orientation="h",
                color="categoria_cor",
                text="Eventos",
                color_discrete_sequence=PALETA_BARRAS
            )

            fig_cpr.update_layout(
                showlegend=False,
                xaxis_title="Eventos",
                yaxis_title="CPR"
            )
            fig_cpr.update_xaxes(tickformat=",d")
            fig_cpr = aplicar_estilo(fig_cpr)
            st.plotly_chart(fig_cpr, use_container_width=True, config={"locale": "pt-BR"})

    # =====================================================
    # EVENTOS POR UPM - FIXO 2026
    # =====================================================

    if coluna_upm:
        st.subheader("🏢 EVENTOS POR UPM")

        upm_df = (
            df_2026[df_2026[coluna_upm].notna()]
            .groupby(coluna_upm)["_ID_LINHA_EVENTO_"]
            .count()
            .reset_index(name="Eventos")
            .sort_values(by="Eventos", ascending=True)
        )

        if not upm_df.empty:
            altura_upm = max(500, len(upm_df) * 45)

            fig_upm = px.bar(
                upm_df,
                x="Eventos",
                y=coluna_upm,
                orientation="h",
                text="Eventos"
            )

            fig_upm.update_traces(
                textposition="outside",
                marker_color=COR_PADRAO_BARRA
            )

            fig_upm = aplicar_estilo(fig_upm)

            fig_upm.update_layout(
                height=altura_upm,
                showlegend=False,
                xaxis_title="Eventos",
                yaxis_title="UPM",
                margin=dict(l=140, r=40, t=60, b=40),
                yaxis=dict(
                    automargin=True,
                    tickfont=dict(size=12),
                    categoryorder="total ascending"
                )
            )

            fig_upm.update_xaxes(tickformat=",d")
            st.plotly_chart(fig_upm, use_container_width=True, config={"locale": "pt-BR"})

    # =====================================================
    # EVENTOS POR NATUREZA - FIXO 2026
    # =====================================================

    if coluna_natureza:
        st.subheader("🎭 EVENTOS POR NATUREZA")

        base_natureza = df_2026[df_2026[coluna_natureza].notna()].copy()

        natureza_df = (
            base_natureza.groupby(coluna_natureza)["_ID_LINHA_EVENTO_"]
            .count()
            .reset_index(name="Eventos")
            .sort_values(by="Eventos", ascending=False)
        )

        if not natureza_df.empty:
            fig_nat = px.pie(
                natureza_df,
                names=coluna_natureza,
                values="Eventos",
                hole=0.45,
                color_discrete_sequence=PALETA_PIZZA
            )
            fig_nat.update_traces(
                textinfo="percent+label",
                textposition="outside",
                hovertemplate="<b>%{label}</b><br>Eventos: %{value}<br>Percentual: %{percent}<extra></extra>"
            )
            fig_nat = aplicar_estilo(fig_nat)
            st.plotly_chart(fig_nat, use_container_width=True, config={"locale": "pt-BR"})

    # =====================================================
    # TRÂNSITO - COLUNA M
    # =====================================================

    st.subheader("🚦 INFORMAÇÕES DE TRÂNSITO")

    if len(df_2026.columns) > 12:
        coluna_transito = df_2026.columns[12]

        base_transito = df_2026.copy()

        base_transito[coluna_transito] = (
            base_transito[coluna_transito]
            .astype("string")
            .fillna("NÃO INFORMADO")
            .str.strip()
            .replace("", "NÃO INFORMADO")
        )

        transito_df = (
            base_transito.groupby(coluna_transito)["_ID_LINHA_EVENTO_"]
            .count()
            .reset_index(name="Eventos")
            .sort_values(by="Eventos", ascending=False)
        )

        if not transito_df.empty:
            fig_transito = px.pie(
                transito_df,
                names=coluna_transito,
                values="Eventos",
                hole=0.45,
                color_discrete_sequence=PALETA_PIZZA
            )

            fig_transito.update_traces(
                textinfo="percent+label",
                textposition="outside",
                hovertemplate="<b>%{label}</b><br>Eventos: %{value}<br>Percentual: %{percent}<extra></extra>"
            )

            fig_transito = aplicar_estilo(fig_transito)

            st.plotly_chart(
                fig_transito,
                use_container_width=True,
                config={"locale": "pt-BR"}
            )
    else:
        st.warning("A planilha não possui a coluna M disponível.")

    # =====================================================
    # EVENTOS POR COMANDO - FIXO 2026
    # =====================================================

    if coluna_comando:
        st.subheader("🚔 EVENTOS POR COMANDO")

        comando_df = (
            df_2026.groupby(coluna_comando)["_ID_LINHA_EVENTO_"]
            .count()
            .reset_index(name="Eventos")
            .sort_values(by="Eventos", ascending=False)
        )

        fig = px.bar(
            comando_df,
            x=coluna_comando,
            y="Eventos",
            color=coluna_comando,
            text_auto=True,
            color_discrete_sequence=PALETA_BARRAS
        )
        fig.update_yaxes(tickformat=",d")
        fig = aplicar_estilo(fig)
        st.plotly_chart(fig, use_container_width=True, config={"locale": "pt-BR"})

       # =====================================================
    # HISTOGRAMA - FIXO 2026
    # =====================================================

    if coluna_publico:
        st.subheader("📊 DISTRIBUIÇÃO DO PÚBLICO")

        base_publico = df_2026[df_2026[coluna_publico].notna()].copy()

        if not base_publico.empty:
            fig = px.histogram(
                base_publico,
                x=coluna_publico,
                nbins=20
            )
            fig.update_traces(marker_color=COR_HISTOGRAMA)
            fig.update_yaxes(tickformat=",d")
            fig = aplicar_estilo(fig)
            st.plotly_chart(fig, use_container_width=True, config={"locale": "pt-BR"})

    # =====================================================
    # EVOLUÇÃO DO PÚBLICO - FIXO 2026
    # =====================================================
    if coluna_publico and df_2026["DATA_EVENTO_BASE"].notna().any():
        st.subheader("📈 EVOLUÇÃO DO PÚBLICO (DIÁRIO)")

        evolucao = (
            df_2026.groupby(df_2026["DATA_EVENTO_BASE"].dt.date)[coluna_publico]
            .sum()
            .reset_index()
        )

        evolucao.columns = ["Data", "Publico"]
        evolucao["Data"] = pd.to_datetime(evolucao["Data"])

        # calendário contínuo
        data_min = evolucao["Data"].min()
        data_max = evolucao["Data"].max()

        calendario = pd.DataFrame({
            "Data": pd.date_range(start=data_min, end=data_max, freq="D")
        })

        evolucao = calendario.merge(evolucao, on="Data", how="left").fillna(0)
        evolucao = evolucao.sort_values("Data")

        # linha mais suave (efeito visual melhor)
        evolucao["Publico_smooth"] = evolucao["Publico"].rolling(3, min_periods=1).mean()

        fig = px.line(
            evolucao,
            x="Data",
            y="Publico",
            markers=True
        )

        # linha principal
        fig.update_traces(
            line=dict(color=COR_LINHA_2, width=3),
            marker=dict(size=6)
        )

        # ✅ eixo X inteligente (menos poluição)
        num_dias = (data_max - data_min).days

        if num_dias <= 10:
            dtick_val = "D1"
        elif num_dias <= 20:
            dtick_val = "D2"
        elif num_dias <= 40:
            dtick_val = "D3"
        else:
            dtick_val = "D5"

        fig.update_xaxes(
            tickformat="%d/%m",
            dtick=dtick_val,
            tickangle=30,
            tickfont=dict(size=11),
            rangeslider=dict(visible=True)  # 🔥 zoom interativo
        )

        fig.update_yaxes(tickformat=",d")

        # ✅ tooltip melhor (informação rica)
        fig.update_traces(
            hovertemplate=(
                "<b>Data:</b> %{x|%d/%m/%Y}<br>"
                "<b>Público:</b> %{y:,}<extra></extra>"
            )
        )

        fig.update_layout(
            xaxis_title="Data",
            yaxis_title="Público Previsto",
            hovermode="x unified"  # 👈 melhora leitura ao passar mouse
        )

        fig = aplicar_estilo(fig)

        st.plotly_chart(
            fig,
            use_container_width=True,
            config={"locale": "pt-BR"}
        )

    # =====================================================
    # EVOLUÇÃO DA QUANTIDADE DE EVENTOS POR HORA
    # =====================================================

    if df_2026["Hora"].notna().any():
        st.subheader("⏰ EVOLUÇÃO DA QUANTIDADE DE EVENTOS POR HORA")

        eventos_hora = (
            df_2026[df_2026["Hora"].notna()]
            .groupby("Hora")["_ID_LINHA_EVENTO_"]
            .count()
            .reset_index(name="Quantidade de Eventos")
            .sort_values("Hora")
        )

        horas_base = pd.DataFrame({"Hora": list(range(24))})
        eventos_hora = horas_base.merge(eventos_hora, on="Hora", how="left").fillna(0)
        eventos_hora["Quantidade de Eventos"] = eventos_hora["Quantidade de Eventos"].astype(int)
        eventos_hora["Hora_Label"] = eventos_hora["Hora"].apply(lambda x: f"{int(x):02d}:00")

        fig_eventos_hora = px.line(
            eventos_hora,
            x="Hora_Label",
            y="Quantidade de Eventos",
            markers=True
        )

        fig_eventos_hora.update_traces(line=dict(color="#5B8FF9", width=3))
        fig_eventos_hora.update_yaxes(tickformat=",d")
        fig_eventos_hora.update_layout(
            xaxis_title="Hora do Dia",
            yaxis_title="Quantidade de Eventos"
        )
        fig_eventos_hora = aplicar_estilo(fig_eventos_hora)
        st.plotly_chart(fig_eventos_hora, use_container_width=True, config={"locale": "pt-BR"})

    # =====================================================
    # MAPA DE CALOR - HORÁRIO X PÚBLICO PREVISTO
    # =====================================================

    if coluna_publico and df_2026["DATA_INICIO_BASE"].notna().any():
        st.subheader("🕒 MAPA DE CALOR - HORÁRIO X PÚBLICO PREVISTO")

        base_calor = (
            df_2026[
                df_2026["DATA_INICIO_BASE"].notna() &
                df_2026["DATA_FIM_BASE"].notna() &
                df_2026[coluna_publico].notna()
            ]
            .copy()
        )

        if not base_calor.empty:
            registros_heatmap = []

            for _, row in base_calor.iterrows():
                inicio = row["DATA_INICIO_BASE"]
                fim = row["DATA_FIM_BASE"]
                publico = row[coluna_publico]

                if pd.isna(inicio) or pd.isna(fim) or pd.isna(publico):
                    continue

                if fim < inicio:
                    fim = inicio

                horas_intervalo = pd.date_range(
                    start=inicio.floor("h"),
                    end=fim.floor("h"),
                    freq="h"
                )

                for dt_ref in horas_intervalo:
                    registros_heatmap.append({
                        "Dia_Semana_Num": dt_ref.weekday(),
                        "Dia_Semana": DIAS_SEMANA[dt_ref.weekday()],
                        "Hora": dt_ref.hour,
                        "Valor": float(publico)
                    })

            heatmap_df = pd.DataFrame(registros_heatmap)

            if not heatmap_df.empty:
                dias_ordem = DIAS_SEMANA
                horas_ordem = list(range(24))

                heatmap_resumo = (
                    heatmap_df.groupby(["Dia_Semana_Num", "Dia_Semana", "Hora"])["Valor"]
                    .sum()
                    .reset_index()
                    .sort_values(["Dia_Semana_Num", "Hora"])
                )

                matriz = (
                    heatmap_resumo.pivot(
                        index="Dia_Semana",
                        columns="Hora",
                        values="Valor"
                    )
                    .reindex(index=dias_ordem, columns=horas_ordem)
                    .fillna(0)
                )

                matriz.columns = [f"{int(col):02d}:00" for col in matriz.columns]

                fig_heat = px.imshow(
                    matriz,
                    aspect="auto",
                    color_continuous_scale="YlOrRd",
                    text_auto=".0f"
                )

                fig_heat.update_layout(
                    xaxis_title="Hora do Dia",
                    yaxis_title="Dia da Semana",
                    coloraxis_colorbar_title="Público Previsto"
                )

                fig_heat.update_xaxes(
                    tickmode="array",
                    tickvals=list(range(len(matriz.columns))),
                    ticktext=list(matriz.columns),
                    side="bottom"
                )

                fig_heat = aplicar_estilo(fig_heat)
                st.plotly_chart(fig_heat, use_container_width=True, config={"locale": "pt-BR"})

    # =====================================================
    # MAPA DE CALOR - HORÁRIO X QUANTIDADE DE EVENTOS
    # =====================================================

    if df_2026["DATA_INICIO_BASE"].notna().any():
        st.subheader("🗓️ MAPA DE CALOR - HORÁRIO X QUANTIDADE DE EVENTOS")

        base_calor_eventos = (
            df_2026[
                df_2026["DATA_INICIO_BASE"].notna() &
                df_2026["DATA_FIM_BASE"].notna()
            ]
            .copy()
        )

        if not base_calor_eventos.empty:
            registros_heatmap_eventos = []

            for _, row in base_calor_eventos.iterrows():
                inicio = row["DATA_INICIO_BASE"]
                fim = row["DATA_FIM_BASE"]

                if pd.isna(inicio) or pd.isna(fim):
                    continue

                if fim < inicio:
                    fim = inicio

                horas_intervalo = pd.date_range(
                    start=inicio.floor("h"),
                    end=fim.floor("h"),
                    freq="h"
                )

                for dt_ref in horas_intervalo:
                    registros_heatmap_eventos.append({
                        "Dia_Semana_Num": dt_ref.weekday(),
                        "Dia_Semana": DIAS_SEMANA[dt_ref.weekday()],
                        "Hora": dt_ref.hour,
                        "Quantidade_Eventos": 1
                    })

            heatmap_eventos_df = pd.DataFrame(registros_heatmap_eventos)

            if not heatmap_eventos_df.empty:
                dias_ordem = DIAS_SEMANA
                horas_ordem = list(range(24))

                heatmap_eventos_resumo = (
                    heatmap_eventos_df.groupby(["Dia_Semana_Num", "Dia_Semana", "Hora"])["Quantidade_Eventos"]
                    .sum()
                    .reset_index()
                    .sort_values(["Dia_Semana_Num", "Hora"])
                )

                matriz_eventos = (
                    heatmap_eventos_resumo.pivot(
                        index="Dia_Semana",
                        columns="Hora",
                        values="Quantidade_Eventos"
                    )
                    .reindex(index=dias_ordem, columns=horas_ordem)
                    .fillna(0)
                )

                matriz_eventos.columns = [f"{int(col):02d}:00" for col in matriz_eventos.columns]

                fig_heat_eventos = px.imshow(
                    matriz_eventos,
                    aspect="auto",
                    color_continuous_scale="Blues",
                    text_auto=".0f"
                )

                fig_heat_eventos.update_layout(
                    xaxis_title="Hora do Dia",
                    yaxis_title="Dia da Semana",
                    coloraxis_colorbar_title="Qtd. Eventos"
                )

                fig_heat_eventos.update_xaxes(
                    tickmode="array",
                    tickvals=list(range(len(matriz_eventos.columns))),
                    ticktext=list(matriz_eventos.columns),
                    side="bottom"
                )

                fig_heat_eventos = aplicar_estilo(fig_heat_eventos)
                st.plotly_chart(fig_heat_eventos, use_container_width=True, config={"locale": "pt-BR"})

    # =====================================================
    # TOP 10 - FIXO 2026
    # =====================================================

    if coluna_publico:
        st.subheader("🔥 TOP 10 - PÚBLICO PREVISTO")

        top_publico = (
            df_2026[df_2026[coluna_publico].notna()]
            .sort_values(by=coluna_publico, ascending=False)
            .head(10)
        )

        st.dataframe(top_publico, use_container_width=True)

        # =====================================================
        # ANÁLISE INTELIGENTE - FIXO 2026
        # =====================================================

        st.subheader("🧠 ANÁLISE INTELIGENTE OPERACIONAL")

        if coluna_publico and not df_2026[coluna_publico].dropna().empty:
            maior_publico = df_2026.loc[df_2026[coluna_publico].idxmax()]
            menor_publico = df_2026.loc[df_2026[coluna_publico].idxmin()]
            media_geral = round(df_2026[coluna_publico].mean(), 2)
            mediana = round(df_2026[coluna_publico].median(), 2)

            nome_maior = maior_publico[coluna_evento] if coluna_evento else "N/D"
            nome_menor = menor_publico[coluna_evento] if coluna_evento else "N/D"

            st.info(f'''
    🚨 EVENTO COM MAIOR PÚBLICO PREVISTO:
    {str(nome_maior).upper()} ({int(maior_publico[coluna_publico]):,} PESSOAS)

    ⚠️ EVENTO COM MENOR PÚBLICO PREVISTO:
    {str(nome_menor).upper()} ({int(menor_publico[coluna_publico]):,} PESSOAS)

    📊 MÉDIA GERAL DE PÚBLICO:
    {media_geral:,.0f} PESSOAS

    📈 MEDIANA DE PÚBLICO:
    {mediana:,.0f} PESSOAS

    🎉 TOTAL DE EVENTOS:
    {total_eventos}

    🏙️ TOTAL DE CIDADES:
    {total_cidades}

    🛡️ TOTAL DE CPR:
    {total_cprs}

    🏢 TOTAL DE UPM:
    {total_upms}
    ''')
        else:
            st.info("Não há dados de público válidos para gerar a análise inteligente.")
    # =====================================================
    # RANKING OPERACIONAL - FIXO 2026
    # =====================================================

    if coluna_comando and coluna_publico:
        st.subheader("🏆 RANKING OPERACIONAL DOS COMANDOS")

        ranking = (
            df_2026.groupby(coluna_comando)[coluna_publico]
            .agg(["sum", "mean", "count"])
            .reset_index()
        )

        ranking.columns = [
            "COMANDO",
            "PÚBLICO TOTAL",
            "MÉDIA PÚBLICO",
            "REGISTROS"
        ]

        ranking = ranking.sort_values("PÚBLICO TOTAL", ascending=False)
        st.dataframe(ranking, use_container_width=True)

    # =====================================================
    # EVENTOS POR TIPO DE PÚBLICO - FIXO 2026
    # =====================================================

    if coluna_tipo_publico:
        st.subheader("👥 EVENTOS POR TIPO DE PÚBLICO")

        tipo_publico_df = (
            df_2026[df_2026[coluna_tipo_publico].notna()]
            .groupby(coluna_tipo_publico)["_ID_LINHA_EVENTO_"]
            .count()
            .reset_index(name="Eventos")
            .sort_values(by="Eventos", ascending=False)
        )

        if not tipo_publico_df.empty:
            fig = px.bar(
                tipo_publico_df,
                x=coluna_tipo_publico,
                y="Eventos",
                color=coluna_tipo_publico,
                text_auto=True,
                color_discrete_sequence=PALETA_BARRAS
            )
            fig.update_yaxes(tickformat=",d")
            fig = aplicar_estilo(fig)
            st.plotly_chart(fig, use_container_width=True, config={"locale": "pt-BR"})

    # =====================================================
    # NATUREZA POR COMANDO - FIXO 2026
    # =====================================================

    if coluna_comando and coluna_natureza:
        st.subheader("🗂️ NATUREZA POR COMANDO")

        nat_comando = (
            df_2026[
                df_2026[coluna_natureza].notna() &
                df_2026[coluna_comando].notna()
            ]
            .groupby([coluna_comando, coluna_natureza])["_ID_LINHA_EVENTO_"]
            .count()
            .reset_index(name="Eventos")
        )

        if not nat_comando.empty:
            fig = px.bar(
                nat_comando,
                x=coluna_comando,
                y="Eventos",
                color=coluna_natureza,
                barmode="stack",
                color_discrete_sequence=PALETA_BARRAS
            )
            fig.update_yaxes(tickformat=",d")
            fig = aplicar_estilo(fig)
            st.plotly_chart(fig, use_container_width=True, config={"locale": "pt-BR"})

        # =====================================================
        # TABELA - DADOS DE TRÂNSITO (SOMENTE DETRAN)
        # =====================================================

    st.subheader("🚦 DADOS DE TRÂNSITO - DETRAN")

    if len(df_2026.columns) > 12:
        coluna_transito = df_2026.columns[12]  # Coluna M
        col_a = df_2026.columns[0]  # Coluna A
        col_c = df_2026.columns[2]  # Coluna C
        col_d = df_2026.columns[3]  # Coluna D
        col_g = df_2026.columns[6]  # Coluna G
        col_h = df_2026.columns[7]  # Coluna H
        col_i = df_2026.columns[8]  # Coluna I

        tabela_transito = df_2026.copy()

        tabela_transito[coluna_transito] = (
            tabela_transito[coluna_transito]
            .astype("string")
            .fillna("")
            .str.strip()
        )

        tabela_transito = tabela_transito[
            tabela_transito[coluna_transito].str.contains("DETRAN", case=False, na=False)
        ].copy()

        tabela_transito = tabela_transito[
            [col_a, col_c, col_d, col_g, col_h, col_i]
        ].copy()

        tabela_transito = tabela_transito.rename(columns={
            col_a: "NÚMERO_SEI"
        })

        if tabela_transito.empty:
            st.info("Não foram encontrados registros de trânsito com DETRAN na coluna M.")
        else:
            st.dataframe(
                tabela_transito,
                use_container_width=True,
                hide_index=True
            )
    else:
        st.warning("A planilha não possui a coluna M disponível.")

    # =====================================================
    # TABELA FINAL - FIXO 2026
    # =====================================================

    st.subheader("📄 DADOS OPERACIONAIS")

    pesquisa = st.text_input("🔎 PESQUISAR NA TABELA")
    tabela = df_2026.copy()

    if pesquisa:
        tabela = tabela[
            tabela.astype(str)
            .apply(lambda x: x.str.contains(pesquisa, case=False, na=False))
            .any(axis=1)
        ]

    st.dataframe(
        tabela.style.apply(lambda row: definir_cor_linha(row, coluna_upm), axis=1),
        use_container_width=True,
        height=450
    )

    # =====================================================
    # DOWNLOAD
    # =====================================================

    csv = tabela.to_csv(index=False).encode("utf-8-sig")

    col_dl1, col_dl2 = st.columns(2)

    with col_dl1:
        st.download_button(
            label="⬇️ BAIXAR CSV",
            data=csv,
            file_name="operacao_sao_joao_2026.csv",
            mime="text/csv"
        )

    with col_dl2:
        if OPENPYXL_DISPONIVEL:
            excel_colorido = gerar_excel_colorido(tabela, coluna_upm)
            st.download_button(
                label="⬇️ BAIXAR EXCEL COLORIDO",
                data=excel_colorido,
                file_name="operacao_sao_joao_2026_colorido.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        else:
            st.warning("Download em Excel colorido desativado: openpyxl não instalado.")

    st.success(f"✅ DASHBOARD ATUALIZADO EM {horario}")

except Exception as erro:
    st.error(f"ERRO AO CARREGAR DADOS: {str(erro)}")
